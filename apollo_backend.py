"""Apollo Backend code
"""
import os
import re
from io import BytesIO
import subprocess
import shutil
import logging
from logging.handlers import RotatingFileHandler
import yaml
import pymongo
from flask import Flask, jsonify, request, send_file
from gevent.pywsgi import WSGIServer
from openpyxl import load_workbook, Workbook
from flask_login import LoginManager, UserMixin, login_user,login_required,logout_user,current_user
from pymongo import MongoClient
from bson import ObjectId
import bcrypt
from flask_cors import CORS

app = Flask(__name__)
CORS(app, supports_credentials=True)
# Set the SameSite attribute for session cookies
app.config['SESSION_COOKIE_SAMESITE'] = 'None'

# Get the absolute path of the configuration file
CONFIG_FILE_PATH = os.path.abspath('backend_config.yaml')
#CONFIG_FILE_PATH = 'backend_config.yaml'

# Load configurations from config.yaml
with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as config_file:
    config_data = yaml.safe_load(config_file)

# MongoDB setup
mongo_uri = (
    f"mongodb://{config_data['mongo']['username']}:"
    f"{config_data['mongo']['password']}@"
    f"{config_data['mongo']['host']}:"
    f"{config_data['mongo']['port']}/"
    f"{config_data['mongo']['db_name']}?authSource=admin"
)
#mongo_client = MongoClient(f"mongodb://mongodb:{config_data['mongo']['port']}/")

# Create a MongoClient instance using the connection string
mongo_client = MongoClient(mongo_uri)


db = mongo_client[config_data['mongo']['db_name']]
users_collection = db['user']
admin_collection = db['admin']
scanner_collection = db['scanners']

# Flask Login setup
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Flask secret key setup
app.secret_key = config_data['flask']['secret_key']

# Global variable to store the path of the original Excel file
ORIGINAL_EXCEL_FILE_PATH = config_data['series_file_path']
ORIGINAL_SCANNER_FILE_PATH =config_data['scanner_file_path']

# Get the directory path from the log file path
log_directory = os.path.dirname(config_data['log_file_path'])

# Create the directory if it doesn't exist
if not os.path.exists(log_directory):
    os.makedirs(log_directory)

# Create the log file if it doesn't exist
log_file_path = config_data['log_file_path'] + config_data['log_file']
if not os.path.exists(log_file_path):
    with open(log_file_path, 'w', encoding='utf-8') as f:
        pass  # Creates an empty file

# The following will handle both file and stream handler
LogFormatStr = '[%(asctime)s] - %(name)s@%(processName)s (%(process)d)\
      - %(levelname)s - %(message)s' # pylint: disable=invalid-name
logging.basicConfig(format=LogFormatStr, level=logging.INFO)
formatter = logging.Formatter(LogFormatStr)
fileHandler = RotatingFileHandler(config_data['log_file_path'] \
                                  + config_data['log_file'], maxBytes=20000000,
                                  backupCount=10)
fileHandler.setLevel(logging.INFO)
fileHandler.setFormatter(formatter)

_logger = logging.getLogger("apollo_backend")
_logger.addHandler(fileHandler)

# Function to make a copy of the original Excel file
def backup_excel_file():
    """Function to make a copy of the original file."""
    # global ORIGINAL_EXCEL_FILE_PATH
    backup_path = ORIGINAL_EXCEL_FILE_PATH + '.bak'
    shutil.copyfile(ORIGINAL_EXCEL_FILE_PATH, backup_path)
    return backup_path
# Function to make a copy of the original Scanner List Excel file
def backup_scanner_list():
    """Function to make a copy the original scanner file."""
    # global ORIGINAL_SCANNER_FILE_PATH
    backup_path = ORIGINAL_SCANNER_FILE_PATH + '.bak'
    shutil.copyfile(ORIGINAL_SCANNER_FILE_PATH, backup_path)
    return backup_path
# Function to revert changes by restoring the original Excel file
def revert_changes():
    """Function to revert changes by restoring the original file."""
    global ORIGINAL_EXCEL_FILE_PATH, ORIGINAL_SCANNER_FILE_PATH  #pylint: disable=global-variable-not-assigned
    # Revert changes to series description Excel file
    series_description_backup_path = ORIGINAL_EXCEL_FILE_PATH + '.bak'
    if os.path.exists(series_description_backup_path):
        shutil.copyfile(series_description_backup_path, ORIGINAL_EXCEL_FILE_PATH)
        os.remove(series_description_backup_path)
    # Revert changes to scanner list Excel file
    scanner_list_backup_path = ORIGINAL_SCANNER_FILE_PATH + '.bak'
    if os.path.exists(scanner_list_backup_path):
        shutil.copyfile(scanner_list_backup_path, ORIGINAL_SCANNER_FILE_PATH)
        os.remove(scanner_list_backup_path)

def copy_excel_to_container():
    """
    Function to copy the Excel files to another Docker container.
    """
    # Get Excel file paths and container name from config
    series_excel_file_path = config_data['series_file_path']
    series_container_name = config_data['series_container_name']
    series_container_destination_path = config_data['series_container_path']

    try:
        # Copy the series description Excel file to the target Docker container
        subprocess.run(['docker', 'cp', series_excel_file_path,
                        f'{series_container_name}:{series_container_destination_path}'], check=True)

        # Get the container ID and image name
        series_container_info = subprocess.run(
            ['docker', 'ps', '-af', f'name={series_container_name}',
             '--format', '{{.ID}}, {{.Image}}'],capture_output=True, text=True, check=True)


        series_container_id, series_container_image = \
            series_container_info.stdout.strip().split(', ')    # pylint: disable=unused-variable

        # Commit changes to new container images
        series_image_name = config_data.get('series_image_name', 'scp')

        subprocess.run(['docker', 'commit', series_container_id, \
                        f'{series_image_name}:{config_data["series_image_tag"]}'], check=True)

        # Stop the Docker container
        subprocess.run(['docker', 'stop', series_container_name], check=True)

        # Start the Docker container
        subprocess.run(['docker', 'start', series_container_name], check=True)

        return True
    except subprocess.CalledProcessError as e:     # pylint: disable=invalid-name
        _logger.error(f"Error copying Excel files to container: {e}") #pylint:disable=logging-fstring-interpolation
        return False
def copy_scanner_to_container():
    """
    Function to copy the Excel files to another Docker container.
    """

    scanner_excel_file_path = config_data['scanner_file_path']
    scanner_container_name = config_data['scanner_container_name']
    scanner_container_destination_path = config_data['scanner_container_path']

    try:
        # Copy the scanner Excel file to the target Docker container
        subprocess.run(['docker', 'cp', scanner_excel_file_path,
            f'{scanner_container_name}:{scanner_container_destination_path}'], check=True)

        scanner_container_info = subprocess.run(
            ['docker', 'ps', '-af', f'name={scanner_container_name}',
             '--format','{{.ID}}, {{.Image}}'],capture_output=True, text=True, check=True)
        scanner_container_id, scanner_container_image = \
            scanner_container_info.stdout.strip().split(', ')    # pylint: disable=unused-variable

        scanner_image_name = config_data.get('scanner_image_name', 'ui_back')
        subprocess.run(['docker', 'commit', scanner_container_id, \
                        f'{scanner_image_name}:{config_data["scanner_image_tag"]}'], check=True)

        # Stop the Docker container
        subprocess.run(['docker', 'stop', scanner_container_name], check=True)

        # Start the Docker container
        subprocess.run(['docker', 'start', scanner_container_name], check=True)

        return True
    except subprocess.CalledProcessError as e:      # pylint: disable=invalid-name
        _logger.error(f"Error copying Excel files to container: {e}") #pylint:disable=logging-fstring-interpolation
        return False


class User(UserMixin):
    """
    Represents a user object for authentication purposes.
    """
    def __init__(self, user_id):
        self.id = str(user_id)     # pylint: disable=C0103

@login_manager.user_loader
def load_user(user_id):
    """
    Callback to reload the user object from the user ID stored in the session.
    """
    return User(user_id)

# Helper functions for Excel operations
def read_excel():
    """
    Read data from an Excel file.
    
    Returns:
        A list of dictionaries containing the series descriptions and sequence types.
    """
    excel_file_path = config_data['series_file_path']
    try:
        workbook = load_workbook(excel_file_path, read_only=True)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            series_description, sequence_type = row
            data.append({'series_description': series_description, 'sequence_type': sequence_type})
        return data
    except Exception as e:    # pylint: disable=C0103, disable=W0718
        _logger.error(f"Error reading Excel file: {e}") #pylint:disable=logging-fstring-interpolation
        return []

def write_to_excel(data):
    """
    Write data to an Excel file.
    
    Args:
        data: A list of dictionaries containing the series descriptions and sequence types.
    """
    excel_file_path = config_data['series_file_path']
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Series Description', 'Sequence Type'])
    for row_data in data:
        series_description = row_data['series_description']
        sequence_type = row_data['sequence_type']
        sheet.append([series_description, sequence_type])
    workbook.save(excel_file_path)

def delete_from_excel(series_description, sequence_type):
    """
    Deletes a specific entry from the Excel file.
    Args:
        series_description: The series description of the entry to be deleted.
        sequence_type: The sequence type of the entry to be deleted.
    Returns:
        True if the entry was successfully deleted, False otherwise.
    """
    existing_data = read_excel()

    for idx, row in enumerate(existing_data):
        if row['series_description'] == series_description\
              and row['sequence_type'] == sequence_type:
            existing_data.pop(idx)
            break
    else:
        return False

    write_to_excel(existing_data)
    return True

def is_sequence_type_allowed(sequence_type):
    """
    Checks if a given sequence type is allowed based on the configuration YAML file.
    Args:
        sequence_type: The sequence type to check.
    Returns:
        True if the sequence type is allowed, False otherwise.
    """
    allowed_sequence_types = config_data.get('allowed_sequence_types', [])
    return sequence_type in allowed_sequence_types

# Routes for User authentication
@app.route('/login', methods=['POST'])
def login():
    """
    Endpoint to handle Admin login.
    """
    _json = request.json
    username = _json['username']
    password = _json['password']

    admin = admin_collection.find_one({'username': username})

    if admin and bcrypt.checkpw(password.encode('utf-8'), admin['password'].encode('utf-8')):
        user_id = str(admin['_id'])
        login_user(User(user_id))
        _logger.info(f"Login successful for {username}") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "Login successful"})
    _logger.error("Invalid username or password")
    return jsonify({"message": "Invalid username or password"}), 401


@app.route('/logout')
@login_required
def logout():
    """
    Endpoint to handle users logout.
    """
    logout_user()
    _logger.info("Logout successful")
    return jsonify({"message": "Logout successful"})

# Routes for managing users
@app.route('/user', methods=['GET'])
@login_required
def get_all_users():
    """
    Endpoint to retrieve all users.
    """
    users = users_collection.find().sort('username', pymongo.ASCENDING)
    user_list = [{'_id': str(user['_id']), 'username': user['username']} for user in users]
    return jsonify(user_list)

@app.route('/user', methods=['POST'])
@login_required
def add_user():
    """
    Endpoint to add a new user.
    """
    _json = request.json
    username = _json['username']
    password = _json['password']
    # admin_username = request.args.get('admin_username')
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    # Ensure username and password are between 6 and 20 characters
    if not (6 <= len(username) <= 20 and 6 <= len(password) <= 20):
        _logger.error("Username and password must be between 6 to 20 characters")
        return jsonify({'message': 'Username and password must '
                        'be between 6 and 20 characters'}), 400

    # Check if username already exists
    existing_user = users_collection.find_one({'username': username})
    if existing_user:
        _logger.error(f"username {username} already exists") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "Username already exists"}), 400

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    user_data = {'username': username, 'password': hashed_password.decode('utf-8')}
    result = users_collection.insert_one(user_data)

    if result.inserted_id:
        _logger.info(f"username {username} added successfully by admin {admin_username}") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "User Added Successfully"})
    return not_found()

@app.route('/user/login', methods=['POST'])
@login_required
def user_login():
    """
    Endpoint to user login.
    """
    _json = request.json
    username = _json['username']
    password = _json['password']

    user = users_collection.find_one({'username': username})

    if user:
        stored_password = user['password']
        if isinstance(stored_password, bytes):
            stored_password_bytes = stored_password
        else:
            stored_password_bytes = stored_password.encode('utf-8')

        if bcrypt.checkpw(password.encode('utf-8'), stored_password_bytes):
            user_id = str(user['_id'])
            login_user(User(user_id))
            _logger.info(f"username {username} logged in successfully") #pylint:disable=logging-fstring-interpolation
            return jsonify({"message": "User login successful"})
    _logger.error("Invalid username or password")
    return jsonify({"message": "Invalid username or password"}), 401


@app.route('/users/<users_id>', methods=['GET'])
@login_required
def get_user(users_id):
    """
    Endpoint to retrieve a specific user.
    """
    user = users_collection.find_one({'_id': ObjectId(users_id)})
    if user:
        user_data = {'_id': str(user['_id']), 'username': user['username']}
        return jsonify(user_data)
    return not_found()

@app.route('/users/<users_id>', methods=['PUT'])
@login_required
def update_user(users_id):
    """
    Endpoint to update a specific user.
    """
    _id = ObjectId(users_id)
    _json = request.json
    username = _json['username']
    password = _json['password']

    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    # Ensure username and password are between 6 and 20 characters
    if not (6 <= len(username) <= 20 and 6 <= len(password) <= 20):
        _logger.error("Username and password must be between 6 to 20 characters")
        return jsonify({"message": "Username and password must be "
                         "between 6 and 20 characters"}), 400

    # Check if the new username already exists in the database
    existing_user = users_collection.find_one({'username': username, '_id': {'$ne': _id}})
    if existing_user:
        _logger.error(f"Username {username} already exists") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "Username already exists"}), 400

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    user_data = {'$set': {'username': username, 'password': hashed_password.decode('utf-8')}}
    result = users_collection.update_one({'_id': _id}, user_data)

    if result.modified_count > 0:
        _logger.info(f"username {username} updated successfully by {admin_username}") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "User Updated Successfully"})
    return not_found()

@app.route('/users/<users_id>', methods=['DELETE'])
@login_required
def delete_user(users_id):
    """
    Endpoint to delete a specific user.
    """
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    # Retrieve user details before deletion
    user = users_collection.find_one({'_id': ObjectId(users_id)})
    if not user:
        return jsonify({"message": "User not found"}), 404

    username = user['username']

    result = users_collection.delete_one({'_id': ObjectId(users_id)})
    if result.deleted_count > 0:
        _logger.info(f"username {username} deleted successfully by {admin_username}") #pylint:disable=logging-fstring-interpolation
        return jsonify({"message": "User deleted successfully"})
    return not_found()

def generate_user_excel():
    """
    Generate an Excel file containing usernames from the users collection.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Users'
    sheet.append(['Username'])  # Adding the header

    users = users_collection.find({}, {'username': 1, '_id': 0})
    for user in users:
        sheet.append([user['username']])

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

@app.route('/download/userlist', methods=['GET'])

def download_user_excel():
    """
    Endpoint to download the updated user Excel sheet.
    """
    try:
        excel_stream = generate_user_excel()

        # Return the Excel file as an attachment for download
        return send_file(excel_stream, as_attachment=True,download_name='userlist.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:   # pylint: disable=C0103, disable=W0718
        _logger.error(f"Error generating or downloading user Excel file: {e}") #pylint:disable=logging-fstring-interpolation
        return jsonify({'error': 'Failed to download user Excel file'}), 500


# Routes for managing series descriptions
@app.route('/series_description', methods=['GET'])
@login_required
def get_series_description():
    """
    Retrieves the series descriptions and sequence types from the Excel file.
    Returns:
        A JSON response containing the series descriptions and sequence types.
    """
    data = read_excel()
    return jsonify(data)

@app.route('/series_description', methods=['POST'])
@login_required
def add_series_description():
    """
    Add a new series description and sequence type to the Excel file.

    Returns:
        A JSON response indicating the success or failure of the operation.
    """
    backup_path = backup_excel_file()
    print(f"Series list backed up to: {backup_path}") #pylint:disable=logging-fstring-interpolation
    new_data = request.get_json()
    if 'series_description' not in new_data or 'sequence_type' not in new_data:
        _logger.error("series_description and sequence_type are required")
        return jsonify({'error': 'series_description and sequence_type are required'}), 400

    series_description = new_data['series_description']
    sequence_type = new_data['sequence_type']

    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    # Check if the series description length exceeds the configured limit
    series_description_limit = config_data.get('series_description_character_limit', 64)
    if len(series_description) > series_description_limit:
        return jsonify({'error': f'Series description must be {series_description_limit} '
                        'characters or fewer'}), 400


    # Check if the sequence type is allowed
    if not is_sequence_type_allowed(sequence_type):
        _logger.error(f"Sequence type {sequence_type} is not allowed") #pylint:disable=logging-fstring-interpolation
        return jsonify({'error': 'Sequence type is not allowed'}), 400

    existing_data = read_excel()
    for row in existing_data:
        if row['series_description'] == series_description:
            _logger.error(f"Series {series_description} already exists") #pylint:disable=logging-fstring-interpolation
            return jsonify({'error': 'This series already exists'}), 400

    existing_data.append(new_data)
    write_to_excel(existing_data)

    if copy_excel_to_container():
        _logger.info(f'Series description {series_description} added successfully by admin' #pylint:disable=logging-fstring-interpolation
                     f' {admin_username} and Excel file copied to Docker container')  #pylint:disable=logging-fstring-interpolation
        return jsonify({'message': 'Data added successfully '
                        'and Excel file copied to Docker container'}), 200

    revert_changes()
    return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500


@app.route('/series_description', methods=['DELETE'])
@login_required
def delete_series_description():
    """
    Delete a series description and sequence type from the Excel file.

    Returns:
        A JSON response indicating the success or failure of the operation.
    """
    backup_path = backup_excel_file()
    print(f"Series list backed up to: {backup_path}")
    request_data = request.get_json()
    if 'series_description' not in request_data or 'sequence_type' not in request_data:
        _logger.error("series_description and sequence_type are required")
        return jsonify({'error': 'series_description and sequence_type are required'}), 400
    series_description = request_data['series_description']
    sequence_type = request_data['sequence_type']
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'
    deleted = delete_from_excel(series_description, sequence_type)
    if deleted:
        if copy_excel_to_container():
            _logger.info(f'series description {series_description} deleted successfully by ' #pylint:disable=logging-fstring-interpolation
                         f'{admin_username} and Excel file copied to Docker container') #pylint:disable=logging-fstring-interpolation
            return jsonify({'message': 'Data deleted successfully and '
                            'Excel file copied to Docker container'
                            }), 200
        revert_changes()
        return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500

    return jsonify({'message': 'Data does not match, not deleted'}), 404

@app.route('/series_description', methods=['PUT'])
@login_required
def update_series_description():                               #pylint: disable=too-many-return-statements
    """
    Updates a series description and sequence type in the Excel file.
    Returns:
        A JSON response indicating the success or failure of the operation.
    """
    backup_path = backup_excel_file()
    print(f"Scanner list backed up to: {backup_path}")
    request_data = request.get_json()
    if 'series_description' not in request_data or 'sequence_type' not in request_data:
        return jsonify({'error': 'series_description and sequence_type are required'}), 400
    series_description = request_data['series_description']
    sequence_type = request_data['sequence_type']
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'
    new_series_description = request_data.get('new_series_description', series_description)
    new_sequence_type = request_data.get('new_sequence_type', sequence_type)

    # Check if the series description length exceeds the configured limit
    series_description_limit = config_data.get('series_description_character_limit', 64)
    if len(new_series_description) > series_description_limit:
        return jsonify({'error': f'Series description must be {series_description_limit} '
                        'characters or fewer'}), 400
    if new_series_description == series_description and new_sequence_type == sequence_type:
        _logger.error("New and old series description and sequence type are same")
        return jsonify({'error':'New and old series description and sequence type are same'}), 400
    existing_data = read_excel()

    # Check if the new series description and sequence type already exist
    for row in existing_data:
        if row['series_description'] == new_series_description:
            return jsonify({'error': 'This series already exists'}), 400
    updated_data = []
    found = False
    for row in existing_data:
        if row['series_description'] == series_description \
            and row['sequence_type'] == sequence_type:
            found = True
            updated_data.append({'series_description': new_series_description,\
                                  'sequence_type': new_sequence_type})
        else:
            updated_data.append(row)
    if not found:
        return jsonify({'message': 'Data not found, not updated'}), 404
    write_to_excel(updated_data)
    if copy_excel_to_container():
        _logger.info(f'series description {series_description} updated successfully by admin ' #pylint:disable=logging-fstring-interpolation
                     f'{admin_username} and Excel file copied to Docker container') #pylint:disable=logging-fstring-interpolation
        return jsonify({'message': 'Data updated successfully and '
                        'Excel file copied to Docker container'}), 200
    revert_changes()
    _logger.error('Failed to copy Excel file to Docker container')
    return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500


@app.route('/download/series', methods=['GET'])

def download_series_excel():
    """
    Downloads the updated Excel sheet for series.
    """
    excel_file_path = config_data.get('series_file_path')
    if not excel_file_path:
        return jsonify({'error': 'Series Excel file path not configured'}), 500

    try:
        # Ensure the Excel file exists
        if not os.path.exists(excel_file_path):
            return jsonify({'error': 'Series Excel file not found'}), 404

        # Return the Excel file as an attachment for download
        return send_file(excel_file_path, as_attachment=True)
    except Exception as e:  # pylint: disable=C0103, disable=W0718
        _logger.error(f"Error downloading series Excel file: {e}") #pylint:disable=logging-fstring-interpolation
        return jsonify({'error': 'Failed to download series Excel file'}), 500


# Helper functions for Scanner List operations
def read_scanner_list():
    """
    Reads scanner list from the Excel file.
    Returns:
        A list of dictionaries containing scanner details.
    """
    scanner_file_path = config_data['scanner_file_path']
    try:
        workbook = load_workbook(scanner_file_path, read_only=True)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            scanner_id, scanner_description, modality = row  # Assuming 'Modality' column exists
            data.append({'scanner_id': scanner_id, \
                         'scanner_description': scanner_description, 'modality': modality,
                         'ip_address': None,
                         'ae_title': None})
        return data
    except Exception as e:   # pylint: disable=C0103, disable=W0718
        _logger.error(f"Error reading Scanner List Excel file: {e}") #pylint:disable=logging-fstring-interpolation
        return []

def write_to_scanner_list(data):
    """
    Writes scanner list data to the Excel file.
    Args:
        data: List of dictionaries containing scanner details.
    """
    scanner_file_path = config_data['scanner_file_path']
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['ScannerId', 'ScannerDescription', 'Modality'])
    for row_data in data:
        scanner_id = row_data['scanner_id']
        scanner_description = row_data['scanner_description']
        sheet.append([scanner_id, scanner_description, 'MR'])
    workbook.save(scanner_file_path)

# Helper function to delete a scanner entry from the scanner list Excel file
def delete_scanner_from_excel(scanner_id):
    """
    Deletes a specific scanner entry from the Excel file.
    Args:
        scanner_id: The ID of the scanner entry to be deleted.
    Returns:
        True if the entry was successfully deleted, False otherwise.
    """
    existing_data = read_scanner_list()

    for idx, row in enumerate(existing_data):
        if row['scanner_id'] == scanner_id:
            existing_data.pop(idx)
            break
    else:
        return False

    write_to_scanner_list(existing_data)
    return True

# Routes for managing scanner list
@app.route('/scanner_list', methods=['GET'])
@login_required
def get_scanner_list():
    """
    Retrieves the scanner list from the Excel file.
    Returns:
        A JSON response containing the scanner list.
    """
    data = read_scanner_list()
    sorted_data = sorted(data, key=lambda x: x['scanner_description']) #Sortby scanner description
    # Also fetch data from MongoDB to include IP and AE Title
    mongo_data = list(scanner_collection.find({}, {'_id': 0}))
    for entry in sorted_data:
        for mongo_entry in mongo_data:
            if entry['scanner_id'] == mongo_entry['scanner_id']:
                entry['ip_address'] = mongo_entry.get('ip_address')
                entry['ae_title'] = mongo_entry.get('ae_title')
                break

    return jsonify(sorted_data)


@app.route('/scanner_list', methods=['POST'])
@login_required
def add_scanner():
    """
    Adds a new scanner to the scanner list.
    """
    backup_path = backup_scanner_list()
    print(f"Scanner list backed up to: {backup_path}")
    new_data = request.get_json()
    if 'scanner_id' not in new_data or 'scanner_description' not in new_data:
        return jsonify({'error': 'scanner_id and scanner_description are required'}), 400

    scanner_id = new_data['scanner_id']
    scanner_description = new_data['scanner_description']
    ip_address = new_data.get('ip_address', None)
    ae_title = new_data.get('ae_title', None)

    # Input validation to check for disallowed characters
    disallowed_chars = r'[\"=]'
    if re.search(disallowed_chars, scanner_description):
        return jsonify({'error': 'Scanner description cannot contain \\ , ", or ='}), 400
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    # Check if the scanner description length exceeds the configured limit
    scanner_description_limit = config_data.get('scanner_description_character_limit', 64)
    if len(scanner_description) > scanner_description_limit:
        return jsonify({'error': f'Scanner description must be {scanner_description_limit} '
                        'characters or fewer'}), 400

    existing_data = read_scanner_list()
    for row in existing_data:
        if row['scanner_id'] == scanner_id:
            _logger.error(f"Scanner {scanner_description} already exists") #pylint:disable=logging-fstring-interpolation
            return jsonify({'error': 'This scanner already exists'}), 400

    existing_data.append(new_data)
    # existing_data.append({'scanner_id': scanner_id, 'scanner_description': scanner_description})

    # Add to MongoDB
    scanner_collection.insert_one({
        'scanner_id': scanner_id,
        'scanner_description': scanner_description,
        'modality': 'MR',
        'ip_address': ip_address,
        'ae_title': ae_title
    })

    write_to_scanner_list(existing_data)

    if copy_scanner_to_container():
        _logger.info(f'Scanner {scanner_description} added successfully by admin ' #pylint:disable=logging-fstring-interpolation
                     f'{admin_username} and Excel file copied to Docker container') #pylint:disable=logging-fstring-interpolation
        return jsonify({'message': 'Scanner added successfully and '
                        'Excel file copied to Docker container'}), 200
    revert_changes()
    _logger.error('Failed to copy Excel file to Docker container')
    return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500


@app.route('/scanner_list', methods=['DELETE'])
@login_required
def delete_scanner():
    """
    Deletes a scanner from the scanner list.
    """
    backup_path= backup_scanner_list()
    print(f"Scanner list backed up to: {backup_path}")
    request_data = request.get_json()
    if 'scanner_id' not in request_data:
        return jsonify({'error': 'scanner_id is required'}), 400
    scanner_id = request_data['scanner_id']

    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    deleted = delete_scanner_from_excel(scanner_id)
    if deleted:
        if copy_scanner_to_container():
            _logger.info(f'Scanner ID {scanner_id} deleted successfully by admin {admin_username}' #pylint:disable=logging-fstring-interpolation
                         ' and Excel file copied to Docker container') 
            return jsonify({'message': 'Scanner deleted successfully and '
                            'Excel file copied to Docker container'}), 200

        revert_changes()
        return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500

    return jsonify({'message': 'Scanner does not exist, not deleted'}), 404

@app.route('/scanner_list', methods=['PUT'])
@login_required
def update_scanner():           #pylint: disable=too-many-return-statements
    """
    Updates a scanner in the scanner list.
    """
    backup_path = backup_scanner_list()
    print(f"Scanner list backed up to: {backup_path}")
    request_data = request.get_json()
    if 'scanner_id' not in request_data or 'scanner_description' not in request_data:
        return jsonify({'error': 'scanner_id, scanner_description are required'}), 400

    scanner_id = request_data['scanner_id']
    scanner_description = request_data['scanner_description']
    ip_address = request_data.get('ip_address', None)
    ae_title = request_data.get('ae_title', None)

    # Check if the scanner description length exceeds the configured limit
    scanner_description_limit = config_data.get('scanner_description_character_limit', 64)
    if len(scanner_description) > scanner_description_limit:
        return jsonify({'error': f'Scanner description must be {scanner_description_limit} '
                        'characters or fewer'}), 400

    # Input validation to check for disallowed characters
    disallowed_chars = r'[\"=]'
    if re.search(disallowed_chars, scanner_description):
        return jsonify({'error': 'Scanner description cannot contain \\, ", or ='}), 400
    admin_id = current_user.id

    # Retrieve admin username from the database using the admin ID
    admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
    admin_username = admin['username'] if admin else 'Unknown Admin'

    existing_data = read_scanner_list()
    updated_data = []
    found = False
    for row in existing_data:
        if row['scanner_id'] == scanner_id:
            found = True
            if row['scanner_description'] == scanner_description:
                _logger.error(f'The scanner description {scanner_description} is already the same') #pylint:disable=logging-fstring-interpolation
                return jsonify({'error': 'The scanner description is already the same'}), 400
            updated_data.append({'scanner_id': scanner_id, \
                                 'scanner_description': scanner_description})
        else:
            updated_data.append(row)
    if not found:
        _logger.error(f"{scanner_id} not found, not updated") #pylint:disable=logging-fstring-interpolation
        return jsonify({'message': 'Scanner not found, not updated'}), 404

    # Update in MongoDB
    scanner_collection.update_one(
        {'scanner_id': scanner_id},
        {'$set': {
            'scanner_description': scanner_description,
            'ip_address': ip_address,
            'ae_title': ae_title
        }}
    )

    write_to_scanner_list(updated_data)
    if copy_scanner_to_container():
        _logger.info(f'Scanner {scanner_description} updated successfully by admin ' #pylint:disable=logging-fstring-interpolation
                     f'{admin_username} and Excel file copied to Docker container') #pylint:disable=logging-fstring-interpolation
        return jsonify({'message': 'Scanner updated successfully and '
                        'Excel file copied to Docker container'}), 200

    revert_changes()
    _logger.error("Failed to copy Excel file to Docker container")
    return jsonify({'error': 'Failed to copy Excel file to Docker container'}), 500


@app.route('/download/scanner', methods=['GET'])

def download_scanner_excel():
    """
    Downloads the updated Excel sheet for scanner.
    """
    excel_file_path = config_data.get('scanner_file_path')
    if not excel_file_path:
        return jsonify({'error': 'Scanner Excel file path not configured'}), 500

    try:
        # Ensure the Excel file exists
        if not os.path.exists(excel_file_path):
            return jsonify({'error': 'Scanner Excel file not found'}), 404

        # Return the Excel file as an attachment for download
        return send_file(excel_file_path, as_attachment=True)
    except Exception as e: # pylint: disable=C0103, disable=W0718
        _logger.error(f"Error downloading scanner Excel file: {e}") #pylint:disable=logging-fstring-interpolation
        return jsonify({'error': 'Failed to download scanner Excel file'}), 500



# Error handler for login_required decorator
@login_manager.unauthorized_handler
def unauthorized_handler():
    """
    Error handler for unauthorized access.
    """
    _logger.error("Login to admin first")
    return jsonify({"message": "Login to admin first"}), 401

@app.errorhandler(404)
def not_found(error=None): # pylint: disable=unused-argument
    """
    Error handler for 404 Not Found.
    """
    message = {
        'status': 404,
        'message': 'Not Found: ' + request.url
    }
    return jsonify(message), 404

if __name__ == '__main__':
    host = config_data.get('flask', {}).get('host', '0.0.0.0')
    port = config_data.get('flask', {}).get('port', 5000)
    http = WSGIServer((host, port), app)
    http.serve_forever()
